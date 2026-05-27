import os
import torch
from torchvision import transforms
import warnings
import json
import openpyxl as op
from torchvision import datasets
import sys
import torchvision

sys.path.append(os.path.abspath(os.path.join(__file__, os.pardir)))

from tools.get_sample_predict import get_sample_predict
import tools.file_utils as ut
from tools.GPU_Detecter import GPU_Detect


class ImageOnlyClassifier:
    """纯图像分类器 - 仅使用四视角图像进行物种分类"""
    
    def __init__(self, device=None):
        self.device = device if device else torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
        self.model = None
        self.class_indict = None
        self.test_transform = transforms.Compose([
            transforms.ToTensor(),
            transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
        ])
    
    def load_genus_model(self, weights_path, num_classes=18):
        """加载属级分类模型"""
        self.model = torchvision.models.efficientnet_b3()
        self.model.classifier[1] = torch.nn.Linear(
            self.model.classifier[1].in_features, num_classes
        )
        self.model = self.model.to(self.device)
        self.model.load_state_dict(torch.load(weights_path, map_location="cpu"))
        self.model.eval()
        print(f"Genus model loaded from {weights_path}")
    
    def load_label_mapping(self, json_path):
        """加载类别标签映射"""
        with open(json_path, "r") as f:
            self.class_indict = json.load(f)
    
    def predict_single_individual(self, ind_path):
        """预测单个个体的属类别"""
        if self.model is None:
            raise RuntimeError("模型未加载，请先调用 load_genus_model")
        
        predict_cla = get_sample_predict(ind_path, self.model, self.device, 18)
        if self.class_indict:
            return self.class_indict[str(predict_cla)]
        return predict_cla
    
    def batch_predict(self, test_path):
        """批量预测测试集中所有个体"""
        if self.model is None:
            raise RuntimeError("模型未加载，请先调用 load_genus_model")
        
        genus_data_dict = {}
        for root, dir_list, file_list in os.walk(test_path):
            if len(dir_list) == 0:
                for file_name in file_list:
                    i = 0
                    for a in list(reversed(root.split('/'))):
                        if a == 'data':
                            break
                        else:
                            i += 1
                    genus_name = list(reversed(root.split('/')))[i-2]
                    ind_name = root.split('/')[-1]
                    ind_path = root
                    if genus_name not in genus_data_dict.keys():
                        genus_data_dict[genus_name] = []
                    genus_data_dict[genus_name].append(ind_path)
        
        genus_predict_dict = {}
        for genus_name in genus_data_dict.keys():
            for ind_path in genus_data_dict[genus_name]:
                predict_cla = get_sample_predict(ind_path, self.model, self.device, 18)
                predict_genus = self.class_indict[str(predict_cla)]
                if predict_genus not in genus_predict_dict.keys():
                    genus_predict_dict[predict_genus] = [ind_path]
                else:
                    genus_predict_dict[predict_genus].append(ind_path)
        
        return genus_predict_dict


def main():
    warnings.filterwarnings("ignore")
    
    # 配置参数
    data_dir = os.path.abspath(os.path.join(__file__, os.pardir))
    test_path = os.path.join(data_dir, "data/test")
    weights_path = os.path.join(data_dir, "weights/EfficientNet-B3/best_network.pth")
    label_path = os.path.join(data_dir, "data/genus_labels.json")
    
    # 检查必要文件
    if not os.path.exists(weights_path):
        print(f"错误：权重文件不存在 {weights_path}")
        sys.exit(1)
    if not os.path.exists(label_path):
        print(f"错误：标签文件不存在 {label_path}")
        sys.exit(1)
    if not os.path.exists(test_path):
        print(f"错误：测试数据路径不存在 {test_path}")
        sys.exit(1)
    
    # 初始化分类器
    classifier = ImageOnlyClassifier()
    classifier.load_genus_model(weights_path)
    classifier.load_label_mapping(label_path)
    
    # 加载测试数据集信息
    test_dataset = datasets.ImageFolder(test_path, classifier.test_transform)
    print(f"测试图像数量: {len(test_dataset)}")
    print(f"类别数量: {len(test_dataset.classes)}")
    print(f"类别名称: {test_dataset.classes}")
    
    # 批量预测
    print("\n开始批量预测...")
    genus_predict_dict = classifier.batch_predict(test_path)
    print("属级预测完成")
    
    # 加载物种分类器
    from tools.SpeciesClassfier_ind import SpeceseClassifier
    species_classifier = SpeceseClassifier(classifier.device, data_dir)
    
    # 物种级预测
    print("开始物种级预测...")
    sp_predict_dict = species_classifier.predict(genus_predict_dict)
    
    # 结果统计和保存
    work_file = op.Workbook()
    sheet = work_file.active
    sheet.append(['ind_name', 'label', 'predict_label', 'consequence'])
    true_num, false_num = 0, 0
    
    for ind_path in sp_predict_dict.keys():
        ind_name = ut.text_segmentation(ind_path, '/')[-1]
        # 获取真实标签
        sp_label = None
        for jpg in os.listdir(ind_path):
            if jpg.lower().endswith(('.jpg', '.jpeg', '.png')):
                sp_label = ut.text_segmentation(jpg, '#')[1]
                break
        
        predict_label = sp_predict_dict[ind_path]
        
        # 处理数字标签
        num_label = False
        for s in predict_label:
            if s.isdigit():
                num_label = True
        if not num_label:
            predict_label = ut.text_segmentation(predict_label, ' ')[-1]
        
        # 判断预测结果
        if sp_label == predict_label:
            true_num += 1
            consequence = "True"
        else:
            false_num += 1
            consequence = "False"
        
        sheet.append([ind_name, sp_label, predict_label, consequence])
    
    # 计算各类别准确率
    output_dict = {}
    for i in range(sheet.max_row - 1):
        row = i + 2
        species_name = sheet.cell(row, 2).value
        if species_name is None:
            continue
        
        num_label = False
        for s in str(species_name):
            if s.isdigit():
                num_label = True
        if not num_label:
            ind_name_val = sheet.cell(row, 1).value
            species_name = ut.text_segmentation(str(ind_name_val), ' ')[0] + ' ' + str(species_name)
        
        if species_name not in output_dict:
            output_dict[species_name] = [0, 0]
        
        consequence = sheet.cell(row, 4).value
        if consequence == 'True':
            output_dict[species_name][0] += 1
            true_num += 1
        else:
            output_dict[species_name][1] += 1
            false_num += 1
    
    # 创建结果汇总表
    work_file.create_sheet('consequence')
    sheet2 = work_file['consequence']
    sheet2.append(["class_name", "number", "true_num", "false_num", "class_acc"])
    
    for species_name in output_dict.keys():
        total = output_dict[species_name][0] + output_dict[species_name][1]
        acc = output_dict[species_name][0] / total if total > 0 else 0
        sheet2.append([species_name, total, output_dict[species_name][0], 
                       output_dict[species_name][1], acc])
    
    # 计算总体准确率
    total_acc = true_num / (true_num + false_num) if (true_num + false_num) > 0 else 0
    sheet2.append(['Total_ACC', '', '', '', total_acc])
    
    # 保存结果
    output_path = os.path.join(data_dir, 'docs/predict_image_only_result.xlsx')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    work_file.save(output_path)
    
    print(f"\n预测完成！总体准确率: {total_acc:.4f}")
    print(f"结果已保存至: {output_path}")


if __name__ == "__main__":
    main()