import os
import torch
from torchvision import transforms
import warnings
import json
import openpyxl as op
from torchvision import datasets
import torch.nn.functional as F
import sys
import torchvision
from tools.SpeciesClassfier_ind import SpeceseClassifier
from tools.get_sample_predict import get_sample_predict
import tools.file_utils as ut
from tools.GPU_Detecter import GPU_Detect
import numpy as np
import pandas as pd 

sys.path.append(os.path.abspath(os.path.join(__file__, os.pardir, os.pardir)))

class SizeInfoLoader:
    """尺寸信息加载器"""
    def __init__(self):
        self.view_order = ["#m#l", "#s#l", "#s#v", "#s#d"]
    
    def extract_view_type(self, image_name):
        """从图片名称中提取视角类型"""
        image_name = str(image_name)
        if "#m#l" in image_name:
            return "#m#l"
        elif "#s#l" in image_name:
            return "#s#l"
        elif "#s#v" in image_name:
            return "#s#v"
        elif "#s#d" in image_name:
            return "#s#d"
        else:
            # 默认处理
            return "#s#d"
    
    def load_size_info(self, ind_path):
        """从个体文件夹加载尺寸信息"""
        size_excel_path = os.path.join(ind_path, "size.xlsx")
        if not os.path.exists(size_excel_path):
            print(f"警告: 在 {ind_path} 中未找到size.xlsx文件")
            return None
        
        try:
            df = pd.read_excel(size_excel_path)
            return df
        except Exception as e:
            print(f"读取尺寸Excel文件失败: {e}")
            return None
    
    def get_size_features(self, ind_path):
        """获取尺寸特征向量"""
        df = self.load_size_info(ind_path)
        if df is None:
            return None
        
        # 提取视角特征
        features = []
        for view in self.view_order:
            view_data = df[df["图片名称"].apply(lambda x: self.extract_view_type(x) == view)]
            if len(view_data) > 0:
                # 取该视角的第一条数据
                row = view_data.iloc[0]
                features.extend([
                    row["length"],
                    row["width"],
                    row["length"] / row["width"],  # aspect_ratio
                    row["length"] * row["width"]   # area
                ])
            else:
                # 缺失视角用0填充
                features.extend([0, 0, 0, 0])
        
        return np.array(features)
    
data_dir = os.path.abspath(os.path.join(__file__, os.pardir, os.pardir))+"/Mogera"
species_classfier_path = (
    "./species_classfier"
)
warnings.filterwarnings("ignore")
device = torch.device("cuda:0")
sp_classfier = SpeceseClassifier(device, species_classfier_path)  # init species classfier
print("Species Classfier Initialized Done")

size_loader = SizeInfoLoader()

# data folder path
dataset_dir = f"{data_dir}"+"/data"
test_path = os.path.join(dataset_dir, "test")

test_transform = transforms.Compose(
    [
     transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
    ]
)

# load test dataset
test_dataset = datasets.ImageFolder(test_path, test_transform)
print("Test image", len(test_dataset))
print("class num", len(test_dataset.classes))
print("class name", test_dataset.classes)

# load json
json_file = open(f"{data_dir}/data/genus_labels.json", "r")
class_indict = json.load(json_file)

# load model
model = torchvision.models.efficientnet_b3()
model.classifier[1] = torch.nn.Linear(model.classifier[1].in_features, 18)
model = model.to(device)
weights_path = f"{data_dir}/weights/EfficientNet-B3/best_network.pth"
model.load_state_dict(torch.load(weights_path, map_location="cpu"))
size_model_path = "size_part/Mogera_train/model/MLP_best.pkl"

genus_data_dict = {} # key:genus, value:ind_path
for root, dir_list, file_list in os.walk(test_path):
    if len(dir_list) == 0:
        for file_name in file_list:
            i = 0
            for a in list(reversed(root.split('/'))):
                if a == 'data':
                    break
                else:
                    i += 1
            genus_name = list(reversed(root.split('/')))[i-2]
            ind_name = root.split('/')[-1]
            ind_path = root
            if genus_name not in genus_data_dict.keys():
                genus_data_dict[genus_name] = []
            genus_data_dict[genus_name].append(ind_path)

model.eval()
# predict genus label
genus_predict_dict = {}
for genus_name in genus_data_dict.keys():
    for ind_path in genus_data_dict[genus_name]:
        # 获取该个体的尺寸信息
        size_features = size_loader.get_size_features(ind_path)
        
        # 修改get_sample_predict调用，传递尺寸信息
        predict_cla = get_sample_predict(ind_path, model, device, 18, size_features=size_features,size_model_path=size_model_path)
        predict_genus = class_indict[str(predict_cla)]
        if predict_genus not in genus_predict_dict.keys():
            genus_predict_dict[predict_genus] = [ind_path]
        else:
            genus_predict_dict[predict_genus].append(ind_path)


# print(df)
print("Genus Predict Done")

# predict species label
sp_predict_dict = sp_classfier.predict(genus_predict_dict)
# print(sp_predict_dict)
work_file = op.Workbook()
sheet = work_file.active
sheet.append(['ind_name','label', 'predict_label', 'consequence'])
true_num, false_num = 0, 0
for ind_path in sp_predict_dict.keys():
    ind_name = ut.text_segmentation(ind_path, '/')[-1]
    for jpg in os.listdir(ind_path):
        sp_label = ut.text_segmentation(jpg, '#')[1]
        break
    predict_label = sp_predict_dict[ind_path]
    append_t = [ind_name, sp_label, predict_label]
    num_label = False
    for s in predict_label:
        if s.isdigit():
            num_label = True
    if not num_label:
        predict_label = ut.text_segmentation(predict_label, ' ')[-1]
    # print(predict_label)
    if sp_label == predict_label:
        true_num += 1
        append_t.append("True")
    else:
        false_num += 1
        append_t.append("False")
    sheet.append(append_t)

# sheet.append([])
# sheet.append(['acc', true_num/(true_num+false_num)])

file = work_file
sheet = file.active
true_num = 0
false_num = 0
output_dict = {}
for i in range(sheet.max_row):
    if sheet.cell(i+2, 1).value == None:
        continue
    species_name = sheet.cell(i+2, 2).value
    species_name = sheet.cell(i+2, 2).value
    num_label = False
    for s in species_name:
        if s.isdigit():
            num_label = True
    if not num_label:
        species_name = ut.text_segmentation(sheet.cell(i+2, 1).value, ' ')[0]+' '+species_name

    if species_name not in output_dict.keys():
        output_dict[species_name] = [0, 0]  # [true, false]
    predict_consequence = sheet.cell(i+2, 4).value
    if predict_consequence == 'True':
        output_dict[species_name][0] += 1
        true_num += 1
    else:
        output_dict[species_name][1] += 1
        false_num += 1
file.create_sheet('consequence')
sheet2 = file['consequence']
sheet2.append(["class_name", "number", "true_num","false_num", "class_acc"])
acc_list = []
for species_name in output_dict.keys():
    num = output_dict[species_name][0]+output_dict[species_name][1]
    acc = output_dict[species_name][0]/num
    sheet2.append([
        species_name,
        num,
        output_dict[species_name][0],
        output_dict[species_name][1],
        acc
        ])
    acc_list.append(acc)

totalacc = true_num/(true_num+false_num)
sheet2.append(['Total_ACC', totalacc])
file.save(f'{data_dir}/docs/predict_ind_ToSpecies_B3.xlsx')
